import re
from datetime import date 
import os.path
import openpyxl
from fpdf import FPDF
import csv 
import pandas as pd

class Person:
    def __init__(self, name=None, DOB=None, Phone=None, email=None):
        self.name = name
        self.DOB = DOB
        self.Phone = Phone
        self.email = email
        if(self.name != None):
            self.age = self.calculateAge()

    def display(self):
        print("Name: ", self.name)
        print("DOB: ", self.DOB)
        print("Phone: ", self.Phone)
        print("Email: ", self.email)
        print("Age: ", self.age)

    def validate(self):
        regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
        is_valid = True
        if(len(self.Phone) != 10):
            is_valid = False
            print("Invalid Phone number")
        if(not(len(self.DOB) == 10 and self.DOB[2] == '/' and self.DOB[5] == '/' and int(self.DOB[0:2]) <= 31 and int(self.DOB[3:5]) > 0 and int(self.DOB[3:5]) <= 12 and int(self.DOB[3:5]) > 0 and int(self.DOB[6:10]) > 1900 and int(self.DOB[6:10]) < 2021)):
            is_valid = False
            print("Invalid Date Of Birth")
        if(not(re.fullmatch(regex, email))):
            is_valid = False
            print("Invalid Email")
        return is_valid
    
    def calculateAge(self):
        today = date.today()
        DOB = self.DOB.split('/')
        self.age = today.year - int(DOB[2]) - ((today.month, today.day) < (int(DOB[1]), int(DOB[0])))
        return self.age
    
    def displaySaveMenu(self):
        print("1. Save in txt file")
        print("2. Save in Excel file")
        print("3. Save in PDF file")

    def saveInTxtFile(self, fname):
        try:
            person_data = ""
            with open('PersonDataSave.csv', mode ='r')as file:
                csvFile = csv.reader(file)
                for lines in csvFile:
                    person_data += lines[0].ljust(20) + str(lines[1]).ljust(20) + lines[2].ljust(20) + lines[3].ljust(20) + lines[4].ljust(20) + "\n"
            with open(fname, 'w') as file:
                file.write(person_data)

            file.close()
            return True
            
        except Exception as e:
            print("Exception occured: " + str(e))
            return False

    def saveInExcel(self):
        try:
            wbk_name = 'PersonData.xlsx'        
            wbk = openpyxl.load_workbook(wbk_name)
            wks = wbk['Data']
        
            data = list(csv.reader(open("PersonDataSave.csv")))
            for i in range(0, len(data)):
                for j in range(0, len(data[i])):
                    wks.cell(row=i+1, column=j+1).value = data[i][j]
            wbk.save(wbk_name)
            wbk.close()
            return True 
        except Exception as e:
            print("Exception occured: " + str(e))
            return False

    def saveInPdf(self, fname):
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size = 12)

            with open('PersonDataSave.csv', mode ='r') as file:
                counter = 1
                csvFile = csv.reader(file)
                for lines in csvFile:
                    person_data = lines[0].ljust(20) + lines[1].ljust(20) + lines[2].ljust(20) + lines[3].ljust(20) + lines[4].ljust(20)
                    pdf.cell(200, 10, txt = person_data, ln = counter)
                    counter += 1
                pdf.output(fname)   
                return True 

        except Exception as e:
            print("Exception occured: " + str(e))
            return False 

    def AddPersonDataToCSV(self):
        try:
            fname = "PersonDataSave.csv"
            person_data_dict = {"Name": self.name, "Age": self.age, "DOB": self.DOB, "Phone": self.Phone, "Email": self.email}
            df = pd.DataFrame([person_data_dict])
            
            if(not(os.path.isfile(fname))):
                df.to_csv(fname, index=False)
            else:
                df.to_csv(fname, mode='a', header = False, index=False)
            return True 
        except Exception as e:
            print("Exception occured: " + str(e))
            return False 

    def BulkReadFromExcelToList(self):
        df = pd.read_excel('PersonData.xlsx')
        for i in df.values.tolist():
            print(i , "\n")

while(True):
    print("1. Add Person data \n2. Export file \n3. Bulk Read from Excel file \n4. Exit")
    choice = int(input("Enter your choice: "))

    if(choice == 1):
        name = input("Enter your name: ")
        DOB = input("Enter your DOB (DD/MM/YYYY): ")
        Phone = input("Enter your phone number: ")
        email = input("Enter your email: ")
        Person1 = Person(name, DOB, Phone, email)
        if(Person1.validate()):
            # Person1.display()
            status = Person1.AddPersonDataToCSV()
            if(status):
                print("Successfully added person data!")
    elif(choice == 2):
        Person().displaySaveMenu()
        export_choice = int(input("Enter your Export choice: "))
        if(export_choice == 1):
            status = Person().saveInTxtFile("PersonData.txt")
            if(status):
                print("Successfully exported to text file!")
        elif(export_choice == 2):
            status = Person().saveInExcel()
            if(status):
                print("Successfully exported to Excel file!")
        elif(export_choice == 3):
            status = Person().saveInPdf("PersonDataForPdf.pdf")
            if(status):
                print("Successfully exported to PDF file!")
    elif(choice == 3):
        Person().BulkReadFromExcelToList() 
    elif(choice == 4):
        break
    else:
        print("Invalid choice!")
